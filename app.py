from flask import Flask, render_template, request, jsonify, redirect, url_for, Response, session
import sqlite3
import os
import json
import functools
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'serp-dashboard-secret-key')


# ---------------------------------------------------------------------------
# Password-only Auth
# ---------------------------------------------------------------------------

@app.before_request
def protect():
    expected = os.environ.get('DASHBOARD_PASSWORD', '')
    if not expected:
        return  # no password set — open access
    if request.endpoint in ('login', 'static'):
        return
    if not session.get('authenticated'):
        return redirect(url_for('login', next=request.path))

@app.route('/login', methods=['GET', 'POST'])
def login():
    expected = os.environ.get('DASHBOARD_PASSWORD', '')
    error = None
    if request.method == 'POST':
        if request.form.get('password') == expected:
            session['authenticated'] = True
            return redirect(request.args.get('next') or '/')
        error = 'Incorrect password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


DB_PATH = os.environ.get(
    'DATABASE_PATH',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'serp_dashboard.db')
)


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS keywords (
            id   INTEGER PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            slug TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS weeks (
            id        INTEGER PRIMARY KEY,
            week_date DATE UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS serp_results (
            id         INTEGER PRIMARY KEY,
            keyword_id INTEGER NOT NULL REFERENCES keywords(id),
            week_id    INTEGER NOT NULL REFERENCES weeks(id),
            position   INTEGER NOT NULL,
            url        TEXT NOT NULL,
            title      TEXT,
            snippet    TEXT,
            UNIQUE(keyword_id, week_id, position)
        );

        CREATE TABLE IF NOT EXISTS url_tags (
            id        INTEGER PRIMARY KEY,
            url       TEXT UNIQUE NOT NULL,
            sentiment TEXT NOT NULL DEFAULT "neutral"
                          CHECK(sentiment IN ("positive","negative","neutral")),
            notes     TEXT DEFAULT "",
            owned     INTEGER NOT NULL DEFAULT 0
        );

        CREATE INDEX IF NOT EXISTS idx_sr_kw_week ON serp_results(keyword_id, week_id);
        CREATE INDEX IF NOT EXISTS idx_sr_url     ON serp_results(url);
    ''')
    conn.commit()

    # Migration: add owned column to existing databases
    try:
        conn.execute('ALTER TABLE url_tags ADD COLUMN owned INTEGER NOT NULL DEFAULT 0')
        conn.commit()
    except Exception:
        pass  # column already exists

    conn.close()


# Ensure the directory for the DB file exists (needed when using a Railway volume path like /data)
_db_dir = os.path.dirname(DB_PATH)
if _db_dir:
    try:
        os.makedirs(_db_dir, exist_ok=True)
    except OSError:
        pass  # directory may already exist or be read-only; sqlite will surface the real error

# Run at import time so gunicorn workers always have a valid schema
init_db()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/robots.txt')
def robots():
    return Response('User-agent: *\nDisallow: /\n', mimetype='text/plain')


@app.route('/')
def index():
    conn = get_db()
    keywords = conn.execute('''
        SELECT * FROM keywords ORDER BY CASE name
            WHEN 'Melaleuca' THEN 1
            WHEN 'Melaleuca.com' THEN 2
            WHEN 'Frank VanderSloot' THEN 3
            WHEN 'Melaleuca Products' THEN 4
            WHEN 'Melaleuca Reviews' THEN 5
            WHEN 'The Wellness Company' THEN 6
            WHEN 'Riverbend Ranch' THEN 7
            ELSE 8
        END
    ''').fetchall()
    weeks    = conn.execute('SELECT * FROM weeks ORDER BY week_date DESC').fetchall()
    conn.close()
    return render_template('index.html', keywords=keywords, weeks=weeks)


@app.route('/api/results')
def api_results():
    keyword_id = request.args.get('keyword_id', type=int)
    week_id    = request.args.get('week_id',    type=int)

    if not keyword_id or not week_id:
        return jsonify([])

    conn = get_db()

    # Current week
    rows = conn.execute('''
        SELECT sr.position, sr.url, sr.title, sr.snippet, ut.sentiment, ut.owned
        FROM   serp_results sr
        LEFT JOIN url_tags ut ON sr.url = ut.url
        WHERE  sr.keyword_id = ? AND sr.week_id = ?
        ORDER  BY sr.position
        LIMIT 100
    ''', (keyword_id, week_id)).fetchall()

    # Previous week
    prev = conn.execute('''
        SELECT id FROM weeks
        WHERE  week_date < (SELECT week_date FROM weeks WHERE id = ?)
        ORDER  BY week_date DESC LIMIT 1
    ''', (week_id,)).fetchone()

    prev_pos = {}
    if prev:
        for r in conn.execute('''
            SELECT url, position FROM serp_results
            WHERE keyword_id = ? AND week_id = ?
        ''', (keyword_id, prev['id'])).fetchall():
            prev_pos[r['url']] = r['position']

    # For URLs absent from the previous week, look up their most recent
    # historical position across ALL prior weeks (not just the last one).
    curr_week_date = conn.execute(
        'SELECT week_date FROM weeks WHERE id=?', (week_id,)
    ).fetchone()['week_date']

    not_in_prev = list({r['url'] for r in rows if r['url'] not in prev_pos})
    last_seen = {}
    if not_in_prev:
        placeholders = ','.join(['?'] * len(not_in_prev))
        for hr in conn.execute(f'''
            SELECT url, position, week_date FROM (
                SELECT sr.url, sr.position, w.week_date,
                       ROW_NUMBER() OVER (PARTITION BY sr.url ORDER BY w.week_date DESC) AS rn
                FROM   serp_results sr
                JOIN   weeks w ON sr.week_id = w.id
                WHERE  sr.keyword_id = ? AND sr.url IN ({placeholders})
                  AND  w.week_date < ?
            ) WHERE rn = 1
        ''', [keyword_id] + not_in_prev + [str(curr_week_date)]).fetchall():
            last_seen[hr['url']] = {'pos': hr['position'], 'date': str(hr['week_date'])}

    conn.close()

    seen_urls = set()
    out = []
    for r in rows:
        url      = r['url']
        is_dup   = url in seen_urls
        seen_urls.add(url)

        last_seen_pos  = None
        last_seen_date = None

        if is_dup:
            movement = 'duplicate'
            mv_val   = 0
        elif url not in prev_pos:
            if url in last_seen:
                movement       = 'returned'
                mv_val         = 0
                last_seen_pos  = last_seen[url]['pos']
                last_seen_date = last_seen[url]['date']
            else:
                movement = 'new'
                mv_val   = 0
        else:
            diff = prev_pos[url] - r['position']   # positive = moved up
            if diff == 0:
                movement = 'no_change'
            elif diff > 0:
                movement = f'up_{diff}'
            else:
                movement = f'down_{abs(diff)}'
            mv_val = diff

        out.append({
            'position':       r['position'],
            'url':            url,
            'title':          r['title']   or '',
            'snippet':        r['snippet'] or '',
            'sentiment':      r['sentiment'] or 'neutral',
            'owned':          bool(r['owned']) if r['owned'] is not None else False,
            'movement':       movement,
            'movement_val':   mv_val,
            'is_duplicate':   is_dup,
            'last_seen_pos':  last_seen_pos,
            'last_seen_date': last_seen_date,
        })

    return jsonify(out)


@app.route('/api/history')
def api_history():
    keyword_id = request.args.get('keyword_id', type=int)
    urls       = request.args.getlist('urls[]')

    if not keyword_id:
        return jsonify({})

    conn = get_db()

    if not urls:
        latest = conn.execute(
            'SELECT id FROM weeks ORDER BY week_date DESC LIMIT 1'
        ).fetchone()
        if not latest:
            conn.close()
            return jsonify({})
        top = conn.execute('''
            SELECT DISTINCT url FROM serp_results
            WHERE keyword_id = ? AND week_id = ?
            ORDER BY position LIMIT 10
        ''', (keyword_id, latest['id'])).fetchall()
        urls = [r['url'] for r in top]

    history = {}
    for url in urls:
        rows = conn.execute('''
            SELECT w.week_date, sr.position
            FROM   serp_results sr
            JOIN   weeks w ON sr.week_id = w.id
            WHERE  sr.keyword_id = ? AND sr.url = ?
            ORDER  BY w.week_date
        ''', (keyword_id, url)).fetchall()
        if rows:
            history[url] = [{'date': str(r['week_date']), 'position': r['position']}
                            for r in rows]

    conn.close()
    return jsonify(history)


@app.route('/api/weeks')
def api_weeks():
    conn  = get_db()
    weeks = conn.execute('SELECT * FROM weeks ORDER BY week_date DESC').fetchall()
    conn.close()
    return jsonify([{'id': w['id'], 'week_date': str(w['week_date'])} for w in weeks])


@app.route('/api/tag', methods=['POST'])
def api_tag():
    data      = request.get_json(force=True)
    url       = data.get('url', '').strip()
    sentiment = data.get('sentiment')
    notes     = data.get('notes', '')
    owned     = data.get('owned')   # None means "don't change it"

    if not url:
        return jsonify({'error': 'url required'}), 400
    if sentiment is not None and sentiment not in ('positive', 'negative', 'neutral'):
        return jsonify({'error': 'Invalid sentiment'}), 400

    conn = get_db()

    if sentiment is not None and owned is not None:
        conn.execute('''
            INSERT INTO url_tags (url, sentiment, notes, owned) VALUES (?, ?, ?, ?)
            ON CONFLICT(url) DO UPDATE SET sentiment=excluded.sentiment, notes=excluded.notes, owned=excluded.owned
        ''', (url, sentiment, notes, int(owned)))
    elif sentiment is not None:
        conn.execute('''
            INSERT INTO url_tags (url, sentiment, notes) VALUES (?, ?, ?)
            ON CONFLICT(url) DO UPDATE SET sentiment=excluded.sentiment, notes=excluded.notes
        ''', (url, sentiment, notes))
    elif owned is not None:
        conn.execute('''
            INSERT INTO url_tags (url, sentiment, notes, owned) VALUES (?, "neutral", "", ?)
            ON CONFLICT(url) DO UPDATE SET owned=excluded.owned
        ''', (url, int(owned)))

    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/upload', methods=['POST'])
def api_upload():
    week_date = request.form.get('week_date', '').strip()
    files     = request.files.getlist('csvfiles')

    if not week_date:
        return jsonify({'error': 'week_date is required'}), 400
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files provided'}), 400

    try:
        datetime.strptime(week_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'week_date must be YYYY-MM-DD'}), 400

    from importer import import_csv_uploads
    try:
        count = import_csv_uploads(DB_PATH, files, week_date)
        # Re-scan XLSX reports for any updated color tags
        try:
            from scan_colors import scan_and_tag
            scan_and_tag(DB_PATH, verbose=False)
        except Exception:
            pass
        return jsonify({'ok': True, 'imported': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/compare')
def api_compare():
    """Return side-by-side results for two weeks."""
    keyword_id = request.args.get('keyword_id', type=int)
    week_a_id  = request.args.get('week_a',     type=int)
    week_b_id  = request.args.get('week_b',     type=int)

    if not all([keyword_id, week_a_id, week_b_id]):
        return jsonify({'error': 'keyword_id, week_a, and week_b are required'}), 400

    conn = get_db()

    def get_week_results(week_id):
        rows = conn.execute('''
            SELECT sr.position, sr.url, sr.title, ut.sentiment, ut.owned
            FROM   serp_results sr
            LEFT JOIN url_tags ut ON sr.url = ut.url
            WHERE  sr.keyword_id = ? AND sr.week_id = ?
            ORDER  BY sr.position
        ''', (keyword_id, week_id)).fetchall()
        return {r['url']: {'position': r['position'], 'title': r['title'] or '', 'sentiment': r['sentiment'] or 'neutral', 'owned': bool(r['owned']) if r['owned'] is not None else False}
                for r in rows}

    week_a_date = conn.execute('SELECT week_date FROM weeks WHERE id=?', (week_a_id,)).fetchone()
    week_b_date = conn.execute('SELECT week_date FROM weeks WHERE id=?', (week_b_id,)).fetchone()

    a = get_week_results(week_a_id)
    b = get_week_results(week_b_id)
    conn.close()

    all_urls = list(dict.fromkeys(list(a.keys()) + list(b.keys())))

    rows = []
    for url in all_urls:
        pos_a = a[url]['position'] if url in a else None
        pos_b = b[url]['position'] if url in b else None
        title = (a.get(url) or b.get(url) or {}).get('title', '')
        sentiment = (a.get(url) or b.get(url) or {}).get('sentiment', 'neutral')

        owned = (a.get(url) or b.get(url) or {}).get('owned', False)

        if pos_a is not None and pos_b is not None:
            diff = pos_a - pos_b   # positive = moved up from A to B
        else:
            diff = None

        rows.append({
            'url':       url,
            'title':     title,
            'sentiment': sentiment,
            'owned':     owned,
            'pos_a':     pos_a,
            'pos_b':     pos_b,
            'diff':      diff,
        })

    rows.sort(key=lambda r: (r['pos_b'] is None, r['pos_b'] or 9999))

    return jsonify({
        'week_a': str(week_a_date['week_date']) if week_a_date else '',
        'week_b': str(week_b_date['week_date']) if week_b_date else '',
        'rows':   rows,
    })


@app.route('/api/volatility')
def api_volatility():
    """Return week-over-week average position change per keyword."""
    keyword_id = request.args.get('keyword_id', type=int)

    conn = get_db()

    kw_filter = 'AND sr.keyword_id = ?' if keyword_id else ''
    params = (keyword_id,) if keyword_id else ()

    weeks = conn.execute(
        'SELECT id, week_date FROM weeks ORDER BY week_date'
    ).fetchall()

    keywords = conn.execute('SELECT id, name FROM keywords ORDER BY name').fetchall()

    result = {}
    for kw in keywords:
        if keyword_id and kw['id'] != keyword_id:
            continue
        kw_data = []
        prev_positions = {}
        for i, week in enumerate(weeks):
            curr = {
                r['url']: r['position']
                for r in conn.execute(
                    'SELECT url, position FROM serp_results WHERE keyword_id=? AND week_id=?',
                    (kw['id'], week['id'])
                ).fetchall()
            }
            if prev_positions and curr:
                shared = set(prev_positions) & set(curr)
                if shared:
                    avg_change = sum(abs(curr[u] - prev_positions[u]) for u in shared) / len(shared)
                    kw_data.append({'date': str(week['week_date']), 'volatility': round(avg_change, 2)})
                else:
                    kw_data.append({'date': str(week['week_date']), 'volatility': None})
            prev_positions = curr
        result[kw['name']] = kw_data

    conn.close()
    return jsonify(result)



@app.route('/api/fetch', methods=['POST'])
def api_fetch():
    data      = request.get_json(force=True)
    week_date = data.get('week_date', '').strip()

    if not week_date:
        return jsonify({'error': 'week_date required'}), 400
    try:
        datetime.strptime(week_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'week_date must be YYYY-MM-DD'}), 400

    api_key = os.environ.get('SCALESERP_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'SCALESERP_API_KEY environment variable is not set'}), 500

    from fetcher import fetch_all
    results = fetch_all(week_date, DB_PATH, api_key)

    total  = sum(v['count'] for v in results.values())
    errors = {k: v['error'] for k, v in results.items() if v['error']}

    return jsonify({'ok': True, 'imported': total, 'results': results, 'errors': errors})


@app.route('/api/fetch_status')
def api_fetch_status():
    """Check whether SCALESERP_API_KEY is configured."""
    configured = bool(os.environ.get('SCALESERP_API_KEY', ''))
    return jsonify({'configured': configured})


@app.route('/api/mozcast')
def api_mozcast():
    """Fetch MozCast temperature data."""
    import urllib.request
    import re
    try:
        req = urllib.request.Request(
            'https://moz.com/mozcast/',
            headers={'User-Agent': 'Mozilla/5.0 (compatible; SERP-Dashboard/1.0)'}
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            html = resp.read().decode('utf-8', errors='ignore')

        # Extract temperature data from the page
        # MozCast embeds chart data as JSON in a script tag
        temp_pattern = re.search(r'"temperatures"\s*:\s*(\[.*?\])', html, re.DOTALL)
        date_pattern = re.search(r'"dates"\s*:\s*(\[.*?\])', html, re.DOTALL)

        if temp_pattern and date_pattern:
            import json as _json
            temps = _json.loads(temp_pattern.group(1))
            dates = _json.loads(date_pattern.group(1))
            data = [{'date': d, 'temp': t} for d, t in zip(dates, temps)]
            return jsonify({'ok': True, 'data': data})

        # Fallback: look for the current temperature display
        current = re.search(r'(\d+(?:\.\d+)?)\s*[°&deg;]', html)
        if current:
            return jsonify({'ok': True, 'data': [], 'current': float(current.group(1))})

        return jsonify({'ok': False, 'error': 'Could not parse MozCast data'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/stats')
def api_stats():
    keyword_id = request.args.get('keyword_id', type=int)
    week_id    = request.args.get('week_id',    type=int)

    if not keyword_id or not week_id:
        return jsonify({})

    conn = get_db()

    total = conn.execute(
        'SELECT COUNT(*) FROM serp_results WHERE keyword_id=? AND week_id=?',
        (keyword_id, week_id)
    ).fetchone()[0]

    # Green = positive + neutral (anything not flagged red)
    neg_count = conn.execute('''
        SELECT COUNT(*) FROM serp_results sr
        JOIN url_tags ut ON sr.url = ut.url
        WHERE sr.keyword_id=? AND sr.week_id=? AND ut.sentiment="negative"
    ''', (keyword_id, week_id)).fetchone()[0]

    pos_count = total - neg_count   # green = everything that isn't red

    prev = conn.execute('''
        SELECT id FROM weeks
        WHERE week_date < (SELECT week_date FROM weeks WHERE id=?)
        ORDER BY week_date DESC LIMIT 1
    ''', (week_id,)).fetchone()

    new_count = 0
    if prev:
        prev_urls = set(
            r[0] for r in conn.execute(
                'SELECT url FROM serp_results WHERE keyword_id=? AND week_id=?',
                (keyword_id, prev['id'])
            ).fetchall()
        )
        curr_urls = set(
            r[0] for r in conn.execute(
                'SELECT url FROM serp_results WHERE keyword_id=? AND week_id=?',
                (keyword_id, week_id)
            ).fetchall()
        )
        new_count = len(curr_urls - prev_urls)

    owned_count = conn.execute('''\
        SELECT COUNT(*) FROM serp_results sr
        JOIN url_tags ut ON sr.url = ut.url
        WHERE sr.keyword_id=? AND sr.week_id=? AND ut.owned=1
    ''', (keyword_id, week_id)).fetchone()[0]

    conn.close()
    return jsonify({
        'total':    total,
        'positive': pos_count,
        'negative': neg_count,
        'new':      new_count,
        'owned':    owned_count,
    })


# ---------------------------------------------------------------------------
# Quick Analysis
# ---------------------------------------------------------------------------

@app.route('/api/analyze', methods=['POST'])
def api_analyze():
    data       = request.get_json(force=True)
    keyword_id = int(data.get('keyword_id', 0))
    week_id    = int(data.get('week_id', 0))

    if not keyword_id or not week_id:
        return jsonify({'error': 'keyword_id and week_id required'}), 400

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY is not configured'}), 500

    conn = get_db()

    kw_row   = conn.execute('SELECT name FROM keywords WHERE id=?', (keyword_id,)).fetchone()
    week_row = conn.execute('SELECT week_date FROM weeks WHERE id=?', (week_id,)).fetchone()
    if not kw_row or not week_row:
        conn.close()
        return jsonify({'error': 'Keyword or week not found'}), 404

    keyword_name  = kw_row['name']
    week_date_str = str(week_row['week_date'])

    # Current week results
    rows = conn.execute('''
        SELECT sr.position, sr.url, sr.title, sr.snippet, ut.sentiment
        FROM   serp_results sr
        LEFT JOIN url_tags ut ON sr.url = ut.url
        WHERE  sr.keyword_id=? AND sr.week_id=?
        ORDER  BY sr.position
        LIMIT 40
    ''', (keyword_id, week_id)).fetchall()

    # Previous week
    prev = conn.execute('''
        SELECT id FROM weeks WHERE week_date < ? ORDER BY week_date DESC LIMIT 1
    ''', (week_date_str,)).fetchone()

    prev_pos = {}
    if prev:
        for r in conn.execute(
            'SELECT url, position FROM serp_results WHERE keyword_id=? AND week_id=?',
            (keyword_id, prev['id'])
        ).fetchall():
            prev_pos[r['url']] = r['position']

    # Last seen for URLs not in prev
    not_in_prev = [r['url'] for r in rows if r['url'] not in prev_pos]
    last_seen = {}
    if not_in_prev:
        placeholders = ','.join(['?'] * len(not_in_prev))
        for hr in conn.execute(f'''
            SELECT url, position, week_date FROM (
                SELECT sr.url, sr.position, w.week_date,
                       ROW_NUMBER() OVER (PARTITION BY sr.url ORDER BY w.week_date DESC) AS rn
                FROM   serp_results sr
                JOIN   weeks w ON sr.week_id = w.id
                WHERE  sr.keyword_id=? AND sr.url IN ({placeholders})
                  AND  w.week_date < ?
            ) WHERE rn=1
        ''', [keyword_id] + not_in_prev + [week_date_str]).fetchall():
            last_seen[hr['url']] = {'pos': hr['position'], 'date': str(hr['week_date'])}

    conn.close()

    # Build a concise data summary for Claude
    seen_urls = set()
    lines = []
    result_count = 0
    for r in rows:
        result_count += 1
        url       = r['url']
        is_dup    = url in seen_urls
        seen_urls.add(url)
        title     = r['title'] or ''
        snippet   = r['snippet'] or ''
        sentiment = r['sentiment'] or 'neutral'

        if is_dup:
            mv_label = 'duplicate'
        elif url not in prev_pos:
            if url in last_seen:
                ls = last_seen[url]
                mv_label = f"returned (last seen #{ls['pos']} on {ls['date']})"
            else:
                mv_label = 'NEW'
        else:
            diff = prev_pos[url] - r['position']
            if diff == 0:    mv_label = 'no change'
            elif diff > 0:   mv_label = f'up {diff}'
            else:            mv_label = f'down {abs(diff)}'

        tag = f'[{sentiment.upper()}]' if sentiment == 'negative' else ''
        lines.append(f"  #{result_count} {tag} {url} | {title} | Movement: {mv_label} | Snippet: {snippet[:120]}")

    serp_text = '\n'.join(lines)

    prompt = f"""You are an SEO analyst assistant. Below is the SERP data for the keyword "{keyword_name}" for the week of {week_date_str}.

SERP DATA:
{serp_text}

Write a concise quick analysis (5-8 bullet points max) that I can share with my boss. Cover:
- Any new links appearing this week and what they are
- Any big movement up or down (especially 5+ positions)
- Any links returning after being absent for a while
- Any negative/concerning results in the top positions
- Overall trend or anything notable

Keep it sharp, professional, and scannable. No fluff. Use plain bullet points (•)."""

    import urllib.request as _req
    payload = json.dumps({
        'model':      'claude-sonnet-4-6',
        'max_tokens': 600,
        'messages':   [{'role': 'user', 'content': prompt}],
    }).encode('utf-8')

    req = _req.Request(
        'https://api.anthropic.com/v1/messages',
        data=payload,
        headers={
            'Content-Type':      'application/json',
            'x-api-key':         api_key,
            'anthropic-version': '2023-06-01',
        },
        method='POST',
    )
    with _req.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode('utf-8'))

    analysis = result['content'][0]['text']
    return jsonify({'ok': True, 'analysis': analysis, 'keyword': keyword_name, 'week': week_date_str})


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

# Maps DB keyword name → Excel sheet name (matching original format)
EXPORT_SHEET_NAMES = {
    'Melaleuca':            'Melaleuca',
    'Melaleuca.com':        'Melaleuca.com',
    'Frank VanderSloot':    'Frank VanderSloot',
    'Melaleuca Products':   'Melaleuca Products',
    'Melaleuca Reviews':    'Melaleuca Reviews',
    'The Wellness Company': 'The Wellness Company (TWC)',
    'Riverbend Ranch':      'Riverbend Ranch (Google)',
}

def _movement_text(movement, last_seen_pos=None, last_seen_date=None):
    if movement == 'no_change':  return 'no change'
    if movement == 'new':        return 'new'
    if movement == 'duplicate':  return 'duplicate'
    if movement == 'returned' and last_seen_pos and last_seen_date:
        d = datetime.strptime(last_seen_date, '%Y-%m-%d')
        return f'{last_seen_pos} on {d.strftime("%-m/%-d/%y")}'
    m = __import__('re').match(r'^(up|down)_(\d+)$', movement or '')
    if m:
        return f'{m.group(1)} {m.group(2)}'
    return movement or ''

@app.route('/api/export')
def api_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    week_id = request.args.get('week_id', type=int)
    if not week_id:
        return jsonify({'error': 'week_id required'}), 400

    conn = get_db()
    week_row = conn.execute('SELECT week_date FROM weeks WHERE id=?', (week_id,)).fetchone()
    if not week_row:
        conn.close()
        return jsonify({'error': 'week not found'}), 404
    week_date_str = str(week_row['week_date'])

    keywords = conn.execute('''
        SELECT * FROM keywords ORDER BY CASE name
            WHEN 'Melaleuca' THEN 1
            WHEN 'Melaleuca.com' THEN 2
            WHEN 'Frank VanderSloot' THEN 3
            WHEN 'Melaleuca Products' THEN 4
            WHEN 'Melaleuca Reviews' THEN 5
            WHEN 'The Wellness Company' THEN 6
            WHEN 'Riverbend Ranch' THEN 7
            ELSE 8
        END
    ''').fetchall()

    # Load all url tags (sentiment + owned) up front
    tag_map = {
        r['url']: {'sentiment': r['sentiment'], 'owned': bool(r['owned'])}
        for r in conn.execute('SELECT url, sentiment, owned FROM url_tags').fetchall()
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin       = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font  = Font(bold=True, size=11)
    header_align = Alignment(horizontal='center')
    page_font    = Font(bold=True, size=11)
    page_align   = Alignment(horizontal='center')
    page_fill    = PatternFill('solid', fgColor='D9D9D9')
    green_fill   = PatternFill('solid', fgColor='00B050')
    red_fill     = PatternFill('solid', fgColor='FF0000')
    yellow_font  = Font(size=11, color='FFFF00')
    black_font   = Font(size=11, color='000000')

    def needs_yellow_movement(mv):
        """Yellow font on movement cell for: new, duplicate, returned, or abs value > 10."""
        if mv in ('new', 'duplicate', 'returned'):
            return True
        import re as _re
        m = _re.match(r'^(up|down)_(\d+)$', mv or '')
        return bool(m and int(m.group(2)) > 10)

    for kw in keywords:
        sheet_name = EXPORT_SHEET_NAMES.get(kw['name'], kw['name'])
        ws = wb.create_sheet(title=sheet_name)

        ws.column_dimensions['A'].width = 15.66
        ws.column_dimensions['B'].width = 92.0
        ws.column_dimensions['C'].width = 17.83

        # Header row
        ws.append(['Position This Week', 'Current URL', 'Movement'])
        for cell in ws[1]:
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border

        # Fetch results
        rows = conn.execute('''
            SELECT sr.position, sr.url
            FROM serp_results sr
            WHERE sr.keyword_id=? AND sr.week_id=?
            ORDER BY sr.position
        ''', (kw['id'], week_id)).fetchall()

        # Previous week
        prev = conn.execute('''
            SELECT id FROM weeks
            WHERE week_date < ? ORDER BY week_date DESC LIMIT 1
        ''', (week_date_str,)).fetchone()

        prev_pos = {}
        if prev:
            for r in conn.execute(
                'SELECT url, position FROM serp_results WHERE keyword_id=? AND week_id=?',
                (kw['id'], prev['id'])
            ).fetchall():
                prev_pos[r['url']] = r['position']

        curr_urls = [r['url'] for r in rows]
        not_in_prev = [u for u in curr_urls if u not in prev_pos]
        last_seen = {}
        if not_in_prev:
            placeholders = ','.join(['?'] * len(not_in_prev))
            for hr in conn.execute(f'''
                SELECT url, position, week_date FROM (
                    SELECT sr.url, sr.position, w.week_date,
                           ROW_NUMBER() OVER (PARTITION BY sr.url ORDER BY w.week_date DESC) AS rn
                    FROM   serp_results sr
                    JOIN   weeks w ON sr.week_id = w.id
                    WHERE  sr.keyword_id=? AND sr.url IN ({placeholders})
                      AND  w.week_date < ?
                ) WHERE rn=1
            ''', [kw['id']] + not_in_prev + [week_date_str]).fetchall():
                last_seen[hr['url']] = {'pos': hr['position'], 'date': str(hr['week_date'])}

        seen_urls = set()
        current_page = 0
        result_count = 0

        for r in rows:
            result_count += 1
            page = (result_count - 1) // 10 + 1
            if page != current_page:
                current_page = page
                page_row_num = ws.max_row + 1
                ws.append([None, f'Page {page}', None])
                for cell in ws[page_row_num]:
                    cell.fill      = page_fill
                    cell.alignment = page_align
                    cell.border    = thin_border
                ws[f'B{page_row_num}'].font = page_font

            url    = r['url']
            is_dup = url in seen_urls
            seen_urls.add(url)

            if is_dup:
                mv = 'duplicate'
                ls_pos, ls_date = None, None
            elif url not in prev_pos:
                if url in last_seen:
                    mv      = 'returned'
                    ls_pos  = last_seen[url]['pos']
                    ls_date = last_seen[url]['date']
                else:
                    mv = 'new'
                    ls_pos, ls_date = None, None
            else:
                diff = prev_pos[url] - r['position']
                if diff == 0:   mv = 'no_change'
                elif diff > 0:  mv = f'up_{diff}'
                else:           mv = f'down_{abs(diff)}'
                ls_pos, ls_date = None, None

            mv_text  = _movement_text(mv, ls_pos, ls_date)
            row_num  = ws.max_row + 1
            ws.append([r['position'], url, mv_text])

            tag       = tag_map.get(url, {})
            sentiment = tag.get('sentiment', 'neutral')
            owned     = tag.get('owned', False)

            fill = red_fill if sentiment == 'negative' else green_fill

            # Borders on all 3 cells
            for col in ('A', 'B', 'C'):
                ws[f'{col}{row_num}'].border = thin_border

            # URL cell: fill + yellow font if starred/owned
            ws[f'B{row_num}'].fill = fill
            ws[f'B{row_num}'].font = yellow_font if owned else black_font

            # Movement cell: fill + yellow font if notable movement
            ws[f'C{row_num}'].fill = fill
            ws[f'C{row_num}'].font = yellow_font if needs_yellow_movement(mv) else black_font

    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    d = datetime.strptime(week_date_str, '%Y-%m-%d')
    filename = f'SERP_Report_{d.strftime("%-m-%-d-%y")}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    init_db()

    conn = get_db()
    count = conn.execute('SELECT COUNT(*) FROM serp_results').fetchone()[0]
    conn.close()

    if count == 0:
        print('Database is empty — importing historical data. This may take a minute...', flush=True)
        from importer import import_all_history
        import_all_history(DB_PATH)
    else:
        print(f'Database ready ({count:,} rows).', flush=True)

    port = int(os.environ.get('PORT', 5050))
    app.run(debug=False, host='0.0.0.0', port=port)
