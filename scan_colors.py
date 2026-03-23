"""
Scan all weekly SERP report XLSX files, read row colors from the URL column,
and upsert the results into url_tags as positive (green) or negative (red).

Green = FF00B050   → sentiment: positive
Red   = FFFF0000   → sentiment: negative
No fill / other    → skip (leave existing tag or default neutral)
"""

import os
import glob
import sqlite3
import openpyxl
from openpyxl.styles.colors import COLOR_INDEX

GREEN  = 'FF00B050'
GREEN2 = 'FF92D050'   # lighter green used in The Wellness Company sheets
RED    = 'FFFF0000'
YELLOW = 'FFFFFF00'   # font color on team-owned/influenced URLs

SERP_BASE = '/Volumes/Creative/L/SEO/SERPs'

# -------------------------------------------------------------------------
# Collect all candidate XLSX report files
# -------------------------------------------------------------------------
def find_all_report_files():
    patterns = [
        '**/SERP-report-*.xlsx',
        '**/SERP_Report*.xlsx',
        '**/Melaleuca_SERP*.xlsx',
        '**/Melaleuca SERPs*.xlsx',
        '**/Melaleuca_SERPs*.xlsx',
        '**/Melaleuca SERPS*.xlsx',
        '**/SERPs *.xlsx',
        '**/SERP_*.xlsx',
        '**/SERPs*.xlsx',
    ]
    found = set()
    for p in patterns:
        for f in glob.glob(os.path.join(SERP_BASE, p), recursive=True):
            base = os.path.basename(f)
            if base.startswith('~$'):       continue   # lock file
            if 'comparison' in base.lower(): continue   # skip comparison files
            if 'analysis'   in base.lower(): continue
            if 'tracking'   in base.lower(): continue
            if 'history'    in base.lower(): continue
            if 'sample'     in base.lower(): continue
            found.add(f)
    return sorted(found)   # chronological-ish by path


# -------------------------------------------------------------------------
# Extract URL→color from one XLSX file
# -------------------------------------------------------------------------
def extract_colors(filepath):
    """Returns (sentiment_dict, owned_set).
    sentiment_dict: {url: 'positive'|'negative'}
    owned_set: {url} for URLs with yellow font color (team-owned/influenced)
    """
    tags  = {}
    owned = set()
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f'  SKIP (unreadable): {os.path.basename(filepath)}: {e}')
        return tags, owned

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            url_cell = None
            # Find the URL cell: col B (index 1)
            for cell in row:
                if cell.column == 2:   # col B
                    url_cell = cell
                    break

            if not url_cell:
                continue
            val = str(url_cell.value or '').strip()
            if not val.startswith('http'):
                continue

            # ---- Check fill color (sentiment) ----
            fill = url_cell.fill
            if fill and fill.fgColor:
                ftype = fill.fgColor.type
                if ftype == 'rgb':
                    color = fill.fgColor.rgb
                elif ftype == 'indexed':
                    idx = fill.fgColor.index
                    if idx == 10:
                        color = RED
                    elif idx == 11:
                        color = GREEN
                    else:
                        color = None
                else:
                    color = None

                if color in (GREEN, GREEN2):
                    tags[val] = 'positive'
                elif color == RED:
                    tags[val] = 'negative'

            # ---- Check font color (owned = yellow) ----
            font = url_cell.font
            if font and font.color and font.color.type == 'rgb':
                if font.color.rgb == YELLOW:
                    owned.add(val)

    wb.close()
    return tags, owned


# -------------------------------------------------------------------------
# Main: scan all files, accumulate tags, upsert to DB
# -------------------------------------------------------------------------
def scan_and_tag(db_path, verbose=True):
    files = find_all_report_files()
    print(f'Found {len(files)} report files to scan...')

    # Process oldest→newest so latest file's color wins
    all_tags  = {}   # url → sentiment
    all_owned = set()
    for filepath in files:
        tags, owned = extract_colors(filepath)
        if (tags or owned) and verbose:
            print(f'  {os.path.basename(filepath):55} → {len(tags)} tagged, {len(owned)} owned')
        all_tags.update(tags)   # newer files override
        all_owned.update(owned)

    print(f'\nTotal unique tagged URLs: {len(all_tags)}')
    pos = sum(1 for s in all_tags.values() if s == 'positive')
    neg = sum(1 for s in all_tags.values() if s == 'negative')
    print(f'  Green (positive): {pos}')
    print(f'  Red   (negative): {neg}')
    print(f'  Yellow (owned):   {len(all_owned)}')

    # Upsert sentiment tags
    conn = sqlite3.connect(db_path)
    inserted = updated = 0
    for url, sentiment in all_tags.items():
        existing = conn.execute('SELECT sentiment FROM url_tags WHERE url=?', (url,)).fetchone()
        if existing:
            if existing[0] != sentiment:
                conn.execute('UPDATE url_tags SET sentiment=? WHERE url=?', (sentiment, url))
                updated += 1
        else:
            conn.execute(
                'INSERT INTO url_tags (url, sentiment, notes) VALUES (?,?,"") ON CONFLICT(url) DO NOTHING',
                (url, sentiment)
            )
            inserted += 1

    # Upsert owned flag (additive — doesn't clear existing sentiment)
    owned_updated = 0
    for url in all_owned:
        existing = conn.execute('SELECT id FROM url_tags WHERE url=?', (url,)).fetchone()
        if existing:
            conn.execute('UPDATE url_tags SET owned=1 WHERE url=?', (url,))
        else:
            conn.execute(
                'INSERT INTO url_tags (url, sentiment, notes, owned) VALUES (?,?,?,1)',
                (url, 'neutral', '')
            )
        owned_updated += 1

    conn.commit()
    conn.close()
    print(f'\nDB updated: {inserted} inserted, {updated} updated, {owned_updated} owned flagged')
    return all_tags


if __name__ == '__main__':
    import sys
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'serp_dashboard.db')
    scan_and_tag(db_path)
